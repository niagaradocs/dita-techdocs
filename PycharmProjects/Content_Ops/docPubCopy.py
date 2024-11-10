#  DO NOT USE - USE 'docPubCopyList.py' instead. Hashtags are breaking the copy process. 10/30/2024
#  This script is for exporting files from the github repo based on the bookmap
#  It copies the bookmap and all child topics into the target folder for exporting to Heretto
#  13 June 2024. Script OK. Added function to also copy the reuse folder so now all wh topics are copied every time.
import os
import re
import shutil
import openpyxl

def normalize_path(file_path, base_path):
    """Normalize paths by resolving any relative references ('../')."""
    # Check if the path is absolute
    if os.path.isabs(file_path):
        return file_path

    # Create a full path if the path is not absolute
    return os.path.normpath(os.path.join(base_path, file_path))

def extract_referenced_files(bookmap_path, base_path, files_found=None):
    if files_found is None:
        files_found = set()

    normalized_path = normalize_path(bookmap_path, base_path)

    with open(normalized_path, 'r', encoding='utf-8') as file:
        content = file.read()

    file_refs = re.findall(r'href="([^"]+)"', content)
    for ref in file_refs:
        normalized_ref = normalize_path(ref, os.path.dirname(normalized_path))
        if normalized_ref.endswith('.ditamap') and normalized_ref not in files_found:
            extract_referenced_files(normalized_ref, base_path, files_found)
        files_found.add(normalized_ref)

    return files_found

def extract_image_references_from_file(xml_path):
    with open(xml_path, "r", encoding='utf-8') as xml_file:
        xml_content = xml_file.read()
    # This regex now looks for the 'image' tag regardless of its parent XML structure.
    return re.findall(r'<image\s+href="([^"]+)"', xml_content)

def find_referenced_graphics(source_root, referenced_files):
    referenced_graphics = set()
    for ref_file in referenced_files:
        ref_path = os.path.join(source_root, ref_file)
        if os.path.exists(ref_path) and ref_path.lower().endswith((".xml", ".dita", ".ditamap")):
            image_references = extract_image_references_from_file(ref_path)
            # Normalize each graphic path
            for image_ref in image_references:
                # Correctly handle relative paths
                image_full_path = normalize_path(image_ref, os.path.dirname(ref_path))
                referenced_graphics.add(image_full_path)
    return referenced_graphics

def write_to_excel(referenced_files, target_folder, docName):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Filename'
    ws['B1'] = 'Folder Name'
    for row, file_path in enumerate(referenced_files, start=2):
        folder_name = os.path.dirname(file_path)
        file_name = os.path.basename(file_path)
        ws[f'A{row}'] = file_name
        ws[f'B{row}'] = folder_name
    excel_path = os.path.join(target_folder, f"{docName}_referenced_files.xlsx")
    wb.save(excel_path)
    print(f"Excel file created at {excel_path}")

def copy_files(referenced_files, source_root, target_root):
    copied_files_count = 0
    for file_path in referenced_files:
        normalized_source_path = normalize_path(file_path, source_root)
        relative_path = os.path.relpath(normalized_source_path, source_root)
        normalized_target_path = os.path.join(target_root, relative_path)
        if not os.path.exists(normalized_source_path):
            print(f"Source file does not exist: {normalized_source_path}")
            continue
        os.makedirs(os.path.dirname(normalized_target_path), exist_ok=True)
        shutil.copy2(normalized_source_path, normalized_target_path)
        copied_files_count += 1
        print(f"Copied and potentially overwritten: {normalized_target_path}")
    return copied_files_count

def main():
    docName = input("What is the Publication name you want to copy? ")
    source_root = "c:\\repos\\dita-niagara"
    target_root = os.path.join("C:\\_publication", docName)
    if not os.path.exists(target_root):
        os.makedirs(target_root)
    bookmap_file_name = f"{docName}.ditamap"
    bookmap_path = os.path.join(source_root, bookmap_file_name)
    referenced_files = extract_referenced_files(bookmap_path, source_root)
    referenced_graphics = find_referenced_graphics(source_root, referenced_files)
    all_referenced_files = {bookmap_file_name}.union(referenced_files, referenced_graphics)
    write_to_excel(all_referenced_files, target_root, docName)
    print(f"Number of files to copy: {len(all_referenced_files)}")
    confirm = input("Do you want to proceed to copy the files to folders? (y/n): ")
    if confirm.lower() == 'y':
        copied_files_count = copy_files(all_referenced_files, source_root, target_root)
        print(f"Copy completed. {copied_files_count} files were copied.")
    else:
        print("Copy operation aborted.")

if __name__ == "__main__":
    main()
